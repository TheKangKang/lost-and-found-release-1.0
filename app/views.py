import hashlib
import os
import uuid

import openpyxl
from django.shortcuts import render

from app.models import User

from django.db.models import Q
from django.http import JsonResponse
import json

from django.conf import settings


# Create your views here.
# 展示用户
def get_users(request):
    # 获取所有普通用户的信息

    try:
        obj_users = User.objects.all().values()
        # 把结果转化为list
        users = list(obj_users)
        return JsonResponse({'code': 1, 'data': users})
    except Exception as e:
        # 如果出现异常，返回
        return JsonResponse({'code': 0, 'msg': "获取普通用户信息出现异常，具体错误：" + str(e)})


# 添加用户
def add_user(request):
    data = json.loads(request.body.decode("utf-8"))

    #  添加
    try:
        obj_user = User(username=data['username'], password=data['password'],
                        real_name=data['real_name'], gender=data['gender'], mobile=data['mobile'],
                        email=data['email'], address=data['address'], image=data['image'], )

        cnt = User.objects.filter(username=data['username']).count()
        if cnt > 0:
            return JsonResponse({'code': 2, 'msg': "用户已存在！"})

        #  执行添加
        obj_user.save()
        try:
            obj_users = User.objects.all().values()
            # 把结果转化为list
            users = list(obj_users)
            return JsonResponse({'code': 1, 'data': users})
        except Exception as e:
            # 如果出现异常，返回
            return JsonResponse({'code': 0, 'msg': "添加用户信息出现异常，具体错误：" + str(e)})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "添加到数据库出现异常，具体：" + str(e)})


# 修改用户
def update_user(request):
    data = json.loads(request.body.decode("utf-8"))

    #  添加
    try:
        #  查找到要修改的用户
        obj_user = User.objects.get(username=data['username'])
        #  执行修改
        obj_user.password = data['password']
        obj_user.real_name = data['real_name']
        obj_user.gender = data['gender']
        obj_user.mobile = data['mobile']
        obj_user.email = data['email']
        obj_user.address = data['address']
        obj_user.image = data['image']

        try:
            obj_user.save()
            obj_users = User.objects.all().values()
            # 把结果转化为list
            users = list(obj_users)
            return JsonResponse({'code': 1, 'data': users})
        except Exception as e:
            # 如果出现异常，返回
            return JsonResponse({'code': 0, 'msg': "修改用户信息出现异常，具体错误：" + str(e)})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "修改到数据库出现异常，具体：" + str(e)})


# 用户登录
def login_user(request):
    data = json.loads(request.body.decode('utf-8'))

    try:
        obj_user = User.objects.get(username=data["username"])
        # print(obj_user.username)
        # print(obj_user.password)
        # print(data['password'])
        if obj_user.password == data['password']:
            session_id = data["username"]
            return JsonResponse({'code': 1, 'sessionId': session_id})
        else:
            return JsonResponse({'code': 0, 'msg': '密码错误'})

    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "登录出现异常，具体：" + str(e)})


# 查询
def query_users(request):
    data = json.loads(request.body.decode('utf-8'))

    try:
        obj_users = User.objects.filter(Q(real_name__contains=data['inputstr']) | Q(gender__contains=data['inputstr']) |
                                        Q(mobile__contains=data['inputstr']) | Q(email__contains=data['inputstr']) |
                                        Q(address__contains=data['inputstr'])).values()
        users = list(obj_users)
        return JsonResponse({'code': 1, 'data': users})
    except Exception as e:
        # 异常
        return JsonResponse({'code': 1, 'msg': "查询用户信息失败：" + str(e)})


# 删除
def delete_user(request):
    data = json.loads(request.body.decode("utf-8"))

    #  添加
    try:
        #  查找到要删除的学生
        obj_user = User.objects.get(username=data['username'])

        obj_user.delete()
        try:
            obj_users = User.objects.all().values()
            # 把结果转化为list
            users = list(obj_users)
            return JsonResponse({'code': 1, 'data': users})
        except Exception as e:
            # 如果出现异常，返回
            return JsonResponse({'code': 0, 'msg': "删除用户信息出现异常，具体错误：" + str(e)})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "删除到数据库出现异常，具体：" + str(e)})


def delete_users(request):
    data = json.loads(request.body.decode('utf-8'))

    try:
        for one_user in data['user']:
            obj_user = User.objects.get(username=one_user['username'])
            obj_user.delete()
        # 获取最后结果
        try:
            obj_users = User.objects.all().values()
            # 把结果转化为listtha
            users = list(obj_users)
            return JsonResponse({'code': 1, 'data': users})
        except Exception as e:
            # 如果出现异常，返回
            return JsonResponse({'code': 0, 'msg': "批量删除用户信息出现异常，具体错误：" + str(e)})
    except Exception as e:
        return JsonResponse({'code': 0, 'msg': "批量删除到数据库出现异常，具体：" + str(e)})


#   随机数
def get_random_str():
    uuid_val = uuid.uuid4()
    uuid_str = str(uuid_val).encode('utf-8')
    md5 = hashlib.md5()
    md5.update(uuid_str)
    return md5.hexdigest()


def import_users_excel(request):
    rev_file = request.FILES.get('excel')
    #   判断是否有文件
    if not rev_file:
        return JsonResponse({'code': 0, 'msg': '文件不存在！'})
    #   获得一个唯一的名字
    excel_name = get_random_str()
    #   准备写入的URL
    file_path = os.path.join(settings.MEDIA_ROOT, excel_name + os.path.splitext(str(rev_file))[1])
    # print("fpath:" + file_path)
    try:
        f = open(file_path, 'wb')
        for i in rev_file.chunks():
            f.write(i)
        f.close()
    except Exception as e:
        JsonResponse({'code': 0, 'msg': str(e)})

    ex_users = read_excel(file_path)
    # print(ex_users)
    success = 0
    error = 0
    error_snos = []
    for one_user in ex_users:
        try:
            obj_user = User.objects.create(username=one_user['username'], password=one_user['password'],
                                           real_name=one_user['real_name'],
                                           gender=one_user['gender'], mobile=one_user['mobile'],
                                           email=one_user['email'], address=one_user['address'])
            success += 1
        except:
            error += 1
            error_snos.append(one_user['username'])
    obj_users = User.objects.all().values()
    users = list(obj_users)
    return JsonResponse({'code': 1, 'success': success, 'error': error, 'errors': error_snos, 'data': users})


def export_users_excel(request):
    obj_users = User.objects.all().values()
    users = list(obj_users)
    # 准备写入的路径
    excel_name = get_random_str() + ".xlsx"
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    # 写入
    write_to_excel(users, path)
    return JsonResponse({'code': 1, 'name': excel_name})


def write_to_excel(data: list, path: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'user'
    keys = data[0].keys()
    for index, item in enumerate(data):
        for k, v in enumerate(keys):
            sheet.cell(row=index + 1, column=k + 1, value=str(item[v]))

    workbook.save(path)


def read_excel(path: str):
    """读取excel路径存储为字典"""
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    users = []
    # 准备key
    keys = ['username', 'password', 'real_name', 'gender', 'mobile', 'email', 'address']
    # 遍历
    for row in sheet.rows:
        # 定义一个临时字典
        temp_dict = {}
        for index, cell in enumerate(row):
            temp_dict[keys[index]] = cell.value
        users.append(temp_dict)
    return users


# 图片上传
def upload(request):
    #   接受上传的文件
    rev_file = request.FILES.get('avatar')
    #   判断是否有文件
    if not rev_file:
        return JsonResponse({'code': 0, 'msg': '图片不存在！'})
    #   获得一个唯一的名字
    new_name = get_random_str()
    #   准备写入的URL
    file_path = os.path.join(settings.MEDIA_ROOT, new_name + os.path.splitext(str(rev_file))[1])
    # print("fpath:" + file_path)
    try:
        f = open(file_path, 'wb')
        for i in rev_file.chunks():
            f.write(i)
        f.close()
        return JsonResponse({'code': 1, 'name': new_name + os.path.splitext(str(rev_file))[1]})
    except Exception as e:
        JsonResponse({'code': 0, 'msg': str(e)})